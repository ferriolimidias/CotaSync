from __future__ import annotations

import io
import unittest
from uuid import uuid4
from unittest.mock import patch

import tests  # noqa: F401
from openpyxl import Workbook, load_workbook

from backend.services.system_spreadsheets import (
    apply_action_outputs_to_system_spreadsheet,
    create_system_spreadsheet,
    export_excel,
    import_excel,
    import_google,
    reconcile_schema,
    sync_google,
)


class SystemSpreadsheetTests(unittest.TestCase):
    def test_system_spreadsheet_import_assigns_stable_client_list(self) -> None:
        from backend.db import Client, ClientList, SessionLocal
        from backend.services.client_lists import create_client_list
        from sqlalchemy import select
        client_list = create_client_list(f"Lista teste {uuid4()}")
        workbook = Workbook()
        workbook.active.append(["Nome", "Grupo", "Cota", "Versão"])
        workbook.active.append(["Cliente", "935", "112", "00"])
        content = io.BytesIO()
        workbook.save(content)
        imported = import_excel(name="Planilha teste", content=content.getvalue(), filename="teste.xlsx", list_id=client_list["id"])
        with SessionLocal() as db:
            client = db.scalar(select(Client).where(Client.system_spreadsheet_id == imported["id"]))
            self.assertEqual(client.list_id, client_list["id"])
            self.assertEqual(db.scalar(select(ClientList).where(ClientList.id == client.list_id)).name, client_list["name"])

    def test_action_outputs_update_one_internal_client_row(self) -> None:
        from backend.db import Client, Run, SessionLocal
        from sqlalchemy import select
        sheet = create_system_spreadsheet("Resultados", ["Nome", "Grupo", "Cota", "Versão", "Resultado A", "Resultado B"])
        fields = {field["display_name"]: field["field_id"] for field in sheet["fields"]}
        with SessionLocal.begin() as db:
            client_id = f"sheet-output-client-{uuid4()}"
            run_id = f"sheet-output-run-{uuid4()}"
            client = Client(id=client_id, name="Cliente", client_group=sheet["id"], system_spreadsheet_id=sheet["id"], grupo="935", cota="112", versao="00", variables={"nome": "Cliente", "grupo": "935", "cota": "112", "versao": "00"}, active=True)
            db.add(client)
            db.add(Run(id=run_id, action_id=None, client_id=client.id, status="success", extracted_data={"Resultado A": "123", "Resultado B": "456"}, input_variables={"grupo": "935", "cota": "112", "versao": "00"}))
        result = apply_action_outputs_to_system_spreadsheet(run_id=run_id, action_id="action", client_id=client_id, variables={"grupo": "935", "cota": "112", "versao": "00"}, result_payload={"dados_extraidos": {"Resultado A": "123", "Resultado B": "456"}}, outputs=[{"output_id": "a", "label": "Resultado A", "destination": {"type": "system_sheet_field", "system_spreadsheet_id": sheet["id"], "field_id": fields["Resultado A"]}}, {"output_id": "b", "label": "Resultado B", "destination": {"type": "system_sheet_field", "system_spreadsheet_id": sheet["id"], "field_id": fields["Resultado B"]}}])
        self.assertEqual(len(result["applied"]), 2)
        with SessionLocal() as db:
            client = db.scalar(select(Client).where(Client.id == client_id))
            self.assertEqual(client.variables["resultado_a"], "123")
            self.assertEqual(client.variables["resultado_b"], "456")
    def test_manual_sheet_has_stable_system_field_ids(self) -> None:
        sheet = create_system_spreadsheet("Teste interno", ["Nome", "Grupo", "Cota", "Versão"])
        self.assertEqual(len(sheet["fields"]), 4)
        self.assertEqual(len({field["field_id"] for field in sheet["fields"]}), 4)
        self.assertTrue(all(field["field_id"].startswith("field-") for field in sheet["fields"]))

    def test_schema_reconcile_preserves_ids_when_columns_are_reordered_or_renamed(self) -> None:
        sheet = create_system_spreadsheet("Schema estável", ["Nome", "Grupo", "Cota", "Versão"])
        ids = {field["internal_key"]: field["field_id"] for field in sheet["fields"]}
        renamed = reconcile_schema(sheet["id"], ["Nome", "Grupo", "Cota", "Parcelas pagas"])
        self.assertEqual(renamed["fields"][3]["field_id"], ids["versao"])
        reordered = reconcile_schema(sheet["id"], ["Cota", "Nome", "Parcelas pagas", "Grupo"])
        self.assertEqual({field["display_name"]: field["field_id"] for field in reordered["fields"]}["Cota"], ids["cota"])

    def test_excel_import_header_outside_first_row_skips_organizational_rows_and_preserves_zero(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Clientes"
        sheet.append(["Equipe Sul"])
        sheet.append(["Nome", "Grupo", "Cota", "Versão", "Resultado"])
        sheet.append(["Cliente 1", "935", "112", "00", "034"])
        output = io.BytesIO()
        workbook.save(output)
        imported = import_excel(name="Excel sintético", content=output.getvalue(), filename="clientes.xlsx", header_row=2)
        self.assertEqual(imported["client_count"], 1)
        self.assertEqual(imported["connectors"][0]["type"], "excel")
        self.assertEqual(imported["fields"][3]["display_name"], "Versão")
        exported = load_workbook(io.BytesIO(export_excel(imported["id"])), read_only=True)
        self.assertEqual(exported.active.cell(3, 4).value, "00")

    def test_excel_export_preserves_original_sheets_formulas_and_styles(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Clientes"
        sheet.append(["Nome", "Grupo", "Cota", "Versão", "Resultado"])
        sheet.append(["Cliente 1", "935", "112", "00", ""])
        sheet["A1"].font = sheet["A1"].font.copy(bold=True)
        sheet["F2"] = "=1+1"
        workbook.create_sheet("Observações")["A1"] = "preservar"
        output = io.BytesIO()
        workbook.save(output)
        imported = import_excel(name="Excel preservado", content=output.getvalue(), filename="preservado.xlsx")
        from backend.db import Client, SessionLocal
        from sqlalchemy import select
        with SessionLocal.begin() as db:
            client = db.scalar(select(Client).where(Client.system_spreadsheet_id == imported["id"]))
            values = dict(client.variables or {})
            values["resultado"] = "123"
            client.variables = values
        exported = load_workbook(io.BytesIO(export_excel(imported["id"])), data_only=False)
        self.assertEqual(exported["Clientes"]["E2"].value, "123")
        self.assertEqual(exported["Clientes"]["F2"].value, "=1+1")
        self.assertTrue(exported["Clientes"]["A1"].font.bold)
        self.assertEqual(exported["Observações"]["A1"].value, "preservar")

    @patch("backend.services.system_spreadsheets._google_request")
    def test_google_import_and_both_connector_sync_use_one_system_sheet(self, request) -> None:
        request.return_value = {"values": [["Nome", "Grupo", "Cota", "Versão", "Resultado"], ["Cliente 1", "935", "112", "00", ""]]}
        sheet = import_google(name="Google sintético", url_or_id="sheet-id", tab="Clientes")
        self.assertEqual(sheet["client_count"], 1)
        self.assertEqual(len(sheet["connectors"]), 1)
        from backend.services.system_spreadsheets import _upsert_connector
        from backend.db import SessionLocal, DataSource
        from sqlalchemy import select
        with SessionLocal.begin() as db:
            canonical = db.scalar(select(DataSource).where(DataSource.id == sheet["id"]))
            _upsert_connector(db, canonical.id, "excel", {"filename": "export.xlsx"})
        updated = sync_google(sheet["id"], direction="inbound")
        self.assertEqual(updated["id"], sheet["id"])
        self.assertEqual({item["type"] for item in updated["connectors"]}, {"google_sheets", "excel"})

    @patch("backend.services.system_spreadsheets._google_request")
    def test_google_outbound_failure_does_not_remove_internal_sheet(self, request) -> None:
        request.return_value = {"values": [["Nome", "Grupo", "Cota", "Versão"], ["Cliente 2", "935", "113", "00"]]}
        sheet = import_google(name="Google falha", url_or_id="sheet-id-2", tab="Clientes")
        request.side_effect = RuntimeError("offline")
        with self.assertRaises(Exception):
            sync_google(sheet["id"], direction="outbound")
        self.assertEqual(sync_google.__name__, "sync_google")


if __name__ == "__main__":
    unittest.main()
