"""Canonical system spreadsheets and their Excel/Google connectors."""
from __future__ import annotations

import io
import json
import os
import re
import base64
import time
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from backend.db import Client, ClientList, DataSource, DataSourceField, Run, SessionLocal, SpreadsheetConnector
from backend.services.actions_repository import project_root
from backend.services.client_fields import canonical_client_field_key

SYSTEM_TYPE = "system_spreadsheet"
CONNECTOR_TYPES = {"excel", "google_sheets"}


class SystemSpreadsheetError(ValueError):
    pass


def _key(value: Any) -> str:
    text = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().casefold()).strip("_")
    return text or "field"


def _field_id(sheet_id: str, internal_key: str) -> str:
    import hashlib
    return f"field-{hashlib.sha256(f'{sheet_id}:{internal_key}'.encode()).hexdigest()[:24]}"


def _dump_field(field: DataSourceField) -> dict[str, Any]:
    return {
        "id": field.id,
        "field_id": field.id,
        "system_spreadsheet_id": field.data_source_id,
        "display_name": field.display_name,
        "internal_key": (field.semantic_role or _key(field.display_name)),
        "position": int(str(field.source_column_reference or "column:0").split(":")[-1]),
        "type": field.data_type,
        "semantic_role": field.semantic_role,
        "active": bool(field.active),
    }


def _dump_sheet(sheet: DataSource, fields: list[DataSourceField], connectors: list[SpreadsheetConnector], client_count: int) -> dict[str, Any]:
    configuration = sheet.configuration or {}
    return {
        "id": sheet.id,
        "system_spreadsheet_id": sheet.id,
        "name": sheet.name,
        "active": sheet.status == "active",
        "identity_mapping": (sheet.configuration or {}).get("identity_mapping", {"grupo": None, "cota": None, "versao": None}),
        "default_list_id": configuration.get("default_list_id"),
        "fields": [_dump_field(field) for field in fields if field.active],
        "client_count": client_count,
        "connectors": [
            {"id": item.id, "type": item.connector_type, "status": item.status, "last_synced_at": item.last_synced_at.isoformat() if item.last_synced_at else None, "last_error": item.last_error}
            for item in connectors
        ],
        "last_sync": max((item.last_synced_at for item in connectors if item.last_synced_at), default=None).isoformat() if any(item.last_synced_at for item in connectors) else None,
    }


def _fields(db, sheet_id: str) -> list[DataSourceField]:
    fields = list(db.scalars(select(DataSourceField).where(DataSourceField.data_source_id == sheet_id)))
    return sorted(fields, key=lambda field: (int(str(field.source_column_reference or "column:0").split(":")[-1]), field.id))


def _sheet(db, sheet_id: str) -> DataSource:
    sheet = db.get(DataSource, sheet_id)
    if sheet is None or sheet.source_type != SYSTEM_TYPE:
        raise SystemSpreadsheetError("Planilha do Sistema não encontrada.")
    return sheet


def _ensure_list(db, list_id: str | None, tenant_id: str) -> ClientList:
    if list_id:
        row = db.scalar(select(ClientList).where(ClientList.id == list_id, ClientList.tenant_id == tenant_id, ClientList.active.is_(True)))
        if row is None:
            raise SystemSpreadsheetError("Lista de clientes não encontrada.")
        return row
    row = db.scalar(select(ClientList).where(ClientList.tenant_id == tenant_id, ClientList.name == "Lista Principal", ClientList.active.is_(True)))
    if row is None:
        row = ClientList(id=str(uuid4()), tenant_id=tenant_id, name="Lista Principal", active=True)
        db.add(row)
        db.flush()
    return row


def create_system_spreadsheet(name: str, headers: list[str], *, identity_mapping: dict[str, Any] | None = None, list_id: str | None = None, tenant_id: str = "default") -> dict[str, Any]:
    clean = [str(value).strip() for value in headers if str(value).strip()]
    if not clean:
        raise SystemSpreadsheetError("A planilha precisa possuir pelo menos um campo.")
    with SessionLocal.begin() as db:
        client_list = _ensure_list(db, list_id, tenant_id)
        sheet = DataSource(id=str(uuid4()), name=str(name).strip() or "Planilha de clientes", source_type=SYSTEM_TYPE, status="active", schema_metadata={"headers": clean, "version": 1}, configuration={"tenant_id": tenant_id, "identity_mapping": identity_mapping or {}, "default_list_id": client_list.id})
        db.add(sheet)
        db.flush()
        for position, header in enumerate(clean):
            internal = canonical_client_field_key(header) or _key(header)
            db.add(DataSourceField(id=_field_id(sheet.id, internal), data_source_id=sheet.id, display_name=header, source_column_reference=f"column:{position}", semantic_role=internal, data_type="string"))
        db.flush()
        return _dump_sheet(sheet, _fields(db, sheet.id), [], 0)


def list_system_spreadsheets(*, tenant_id: str = "default") -> list[dict[str, Any]]:
    with SessionLocal() as db:
        sheets = list(db.scalars(select(DataSource).where(DataSource.source_type == SYSTEM_TYPE).order_by(DataSource.created_at.desc())))
        result = []
        for sheet in sheets:
            if (sheet.configuration or {}).get("tenant_id", "default") != tenant_id:
                continue
            connectors = list(db.scalars(select(SpreadsheetConnector).where(SpreadsheetConnector.spreadsheet_id == sheet.id)))
            count = db.query(Client).filter(Client.system_spreadsheet_id == sheet.id).count()
            result.append(_dump_sheet(sheet, _fields(db, sheet.id), connectors, count))
        return result


def get_system_spreadsheet(sheet_id: str, *, tenant_id: str = "default") -> dict[str, Any]:
    with SessionLocal() as db:
        sheet = _sheet(db, sheet_id)
        if (sheet.configuration or {}).get("tenant_id", "default") != tenant_id:
            raise SystemSpreadsheetError("Planilha do Sistema não encontrada.")
        connectors = list(db.scalars(select(SpreadsheetConnector).where(SpreadsheetConnector.spreadsheet_id == sheet.id)))
        count = db.query(Client).filter(Client.system_spreadsheet_id == sheet.id).count()
        return _dump_sheet(sheet, _fields(db, sheet.id), connectors, count)


def get_system_spreadsheet_rows(sheet_id: str, *, tenant_id: str = "default") -> dict[str, Any]:
    sheet = get_system_spreadsheet(sheet_id, tenant_id=tenant_id)
    with SessionLocal() as db:
        clients = list(db.scalars(select(Client).where(Client.system_spreadsheet_id == sheet_id).order_by(Client.name)))
        rows = []
        for client in clients:
            values = dict(client.variables or {})
            rows.append({"client_id": client.id, "name": client.name, "active": client.active, "values": {field["internal_key"]: values.get(field["internal_key"], "") for field in sheet["fields"]}})
    return {"system_spreadsheet": sheet, "rows": rows}


def update_system_spreadsheet_row(sheet_id: str, client_id: str, values: dict[str, Any], *, tenant_id: str = "default") -> dict[str, Any]:
    sheet = get_system_spreadsheet(sheet_id, tenant_id=tenant_id)
    identity_keys = {"grupo", "cota", "versao"}
    if identity_keys.intersection(values):
        raise SystemSpreadsheetError("Campos de identidade devem ser alterados pela reconciliação da planilha.")
    allowed = {field["internal_key"] for field in sheet["fields"]}
    unknown = set(values) - allowed
    if unknown:
        raise SystemSpreadsheetError("A linha contém campos que não pertencem à Planilha do Sistema.")
    with SessionLocal.begin() as db:
        client = db.scalar(select(Client).where(Client.id == client_id, Client.system_spreadsheet_id == sheet_id))
        if client is None:
            raise SystemSpreadsheetError("Cliente não pertence à Planilha do Sistema.")
        merged = dict(client.variables or {})
        merged.update({key: "" if value is None else str(value) for key, value in values.items()})
        client.variables = merged
        for connector in db.scalars(select(SpreadsheetConnector).where(SpreadsheetConnector.spreadsheet_id == sheet_id)):
            connector.status = "pending"
    return get_system_spreadsheet_rows(sheet_id, tenant_id=tenant_id)


def reconcile_schema(sheet_id: str, headers: list[str], *, tenant_id: str = "default") -> dict[str, Any]:
    with SessionLocal.begin() as db:
        sheet = _sheet(db, sheet_id)
        if (sheet.configuration or {}).get("tenant_id", "default") != tenant_id:
            raise SystemSpreadsheetError("Planilha do Sistema não encontrada.")
        _upsert_fields(db, sheet, [str(value).strip() for value in headers if str(value).strip()])
    return get_system_spreadsheet(sheet_id, tenant_id=tenant_id)


def _upsert_connector(db, sheet_id: str, connector_type: str, configuration: dict[str, Any]) -> SpreadsheetConnector:
    connector = db.scalar(select(SpreadsheetConnector).where(SpreadsheetConnector.spreadsheet_id == sheet_id, SpreadsheetConnector.connector_type == connector_type))
    if connector is None:
        connector = SpreadsheetConnector(id=str(uuid4()), spreadsheet_id=sheet_id, connector_type=connector_type)
        db.add(connector)
    connector.configuration = {**(connector.configuration or {}), **configuration}
    connector.status = "pending"
    connector.last_error = None
    return connector


def attach_excel(*, sheet_id: str, content: bytes, filename: str, sheet_name: str | None = None, header_row: int = 1, tenant_id: str = "default") -> dict[str, Any]:
    sheet = get_system_spreadsheet(sheet_id, tenant_id=tenant_id)
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise SystemSpreadsheetError("Arquivo Excel inválido.") from exc
    selected = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(selected.iter_rows(values_only=True))
    if header_row < 1 or header_row > len(rows):
        raise SystemSpreadsheetError("Linha de cabeçalho inválida.")
    headers = [str(value or "").strip() for value in rows[header_row - 1]]
    path = project_root() / "data" / "spreadsheets" / f"{sheet_id}-{re.sub(r'[^A-Za-z0-9_.-]', '_', filename)}"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)
    with SessionLocal.begin() as db:
        canonical = _sheet(db, sheet_id)
        _upsert_fields(db, canonical, headers)
        _upsert_connector(db, sheet_id, "excel", {"filename": filename, "workbook_path": str(path), "sheet_name": selected.title, "header_row": header_row})
        _upsert_rows(db, sheet_id, headers, [list(row) for row in rows[header_row:]])
    return get_system_spreadsheet(sheet_id, tenant_id=tenant_id)


def _upsert_fields(db, sheet: DataSource, headers: list[str]) -> list[DataSourceField]:
    fields = _fields(db, sheet.id)
    by_name = {_key(field.display_name): field for field in fields}
    by_position = {int(str(field.source_column_reference or "column:0").split(":")[-1]): field for field in fields}
    for position, header in enumerate(headers):
        internal = canonical_client_field_key(header) or _key(header)
        field = by_name.get(internal) or by_position.get(position)
        if field is None:
            field = DataSourceField(id=_field_id(sheet.id, internal), data_source_id=sheet.id, display_name=header, source_column_reference=f"column:{position}", semantic_role=internal, data_type="string")
            db.add(field)
        else:
            field.display_name = header
            field.source_column_reference = f"column:{position}"
            field.semantic_role = internal
            field.active = True
    sheet.schema_metadata = {"headers": headers, "version": int((sheet.schema_metadata or {}).get("version", 0)) + 1}
    db.flush()
    return _fields(db, sheet.id)


def _upsert_rows(db, sheet_id: str, headers: list[str], rows: list[list[Any]], identity_mapping: dict[str, Any] | None = None, list_id: str | None = None) -> int:
    fields = _fields(db, sheet_id)
    sheet = _sheet(db, sheet_id)
    configured_list_id = list_id or (sheet.configuration or {}).get("default_list_id")
    client_list = _ensure_list(db, configured_list_id, (sheet.configuration or {}).get("tenant_id", "default"))
    keys = [field.semantic_role or _key(field.display_name) for field in fields]
    identity = identity_mapping or {}
    role_to_key = {role: str(identity.get(role) or role) for role in ("grupo", "cota", "versao")}
    existing = {f"{str(client.grupo or '')}|{str(client.cota or '')}|{str(client.versao or '')}": client for client in db.scalars(select(Client).where(Client.system_spreadsheet_id == sheet_id))}
    count = 0
    for row in rows:
        values = {keys[index]: "" if value is None else str(value) for index, value in enumerate(row[:len(keys)])}
        grupo, cota, versao = (values.get(role_to_key[role], "") for role in ("grupo", "cota", "versao"))
        if not (grupo or cota or versao):
            continue
        if not grupo or not cota:
            continue
        row_key = f"{grupo}|{cota}|{versao}"
        client = existing.get(row_key)
        if client is None:
            client = Client(id=str(uuid4()), name=values.get("name") or values.get("nome") or row_key, client_group=client_list.name, list_id=client_list.id, system_spreadsheet_id=sheet_id, active=True)
            db.add(client)
            existing[row_key] = client
        client.name = values.get("name") or values.get("nome") or client.name
        client.client_group = client_list.name
        client.list_id = client_list.id
        client.system_spreadsheet_id = sheet_id
        client.grupo, client.cota, client.versao = grupo, cota, versao
        client.variables = values
        count += 1
    return count


def import_excel(*, name: str, content: bytes, filename: str, sheet_name: str | None = None, header_row: int = 1, identity_mapping: dict[str, Any] | None = None, list_id: str | None = None, tenant_id: str = "default") -> dict[str, Any]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise SystemSpreadsheetError("Arquivo Excel inválido.") from exc
    selected = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(selected.iter_rows(values_only=True))
    if header_row < 1 or header_row > len(rows):
        raise SystemSpreadsheetError("Linha de cabeçalho inválida.")
    headers = [str(value or "").strip() for value in rows[header_row - 1]]
    if not any(headers):
        raise SystemSpreadsheetError("Não foi possível detectar o cabeçalho.")
    sheet_data = create_system_spreadsheet(name, headers, identity_mapping=identity_mapping, list_id=list_id, tenant_id=tenant_id)
    path = project_root() / "data" / "spreadsheets" / f"{sheet_data['id']}-{re.sub(r'[^A-Za-z0-9_.-]', '_', filename)}"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)
    with SessionLocal.begin() as db:
        sheet = _sheet(db, sheet_data["id"])
        _upsert_connector(db, sheet.id, "excel", {"filename": filename, "workbook_path": str(path), "sheet_name": selected.title, "header_row": header_row})
        _upsert_rows(db, sheet.id, headers, [list(row) for row in rows[header_row:]], identity_mapping, list_id)
    return get_system_spreadsheet(sheet_data["id"], tenant_id=tenant_id)


def export_excel(sheet_id: str, *, tenant_id: str = "default") -> bytes:
    sheet = get_system_spreadsheet(sheet_id, tenant_id=tenant_id)
    fields = sheet["fields"]
    with SessionLocal() as db:
        clients = list(db.scalars(select(Client).where(Client.system_spreadsheet_id == sheet_id).order_by(Client.name)))
        connector = db.scalar(select(SpreadsheetConnector).where(SpreadsheetConnector.spreadsheet_id == sheet_id, SpreadsheetConnector.connector_type == "excel"))
        config = dict(connector.configuration or {}) if connector else {}
    workbook_path = Path(str(config.get("workbook_path") or ""))
    if workbook_path.is_file():
        wb = load_workbook(workbook_path, data_only=False)
        ws = wb[str(config.get("sheet_name") or wb.sheetnames[0])]
        header_row = int(config.get("header_row") or 1)
        headers = {str(ws.cell(header_row, column).value or "").strip(): column for column in range(1, ws.max_column + 1)}
        positions = {field["internal_key"]: headers.get(field["display_name"]) for field in fields}
        identity = {key: positions.get(key) for key in ("grupo", "cota", "versao")}
        existing: dict[str, int] = {}
        for row_number in range(header_row + 1, ws.max_row + 1):
            key = "|".join(str(ws.cell(row_number, identity[key]).value or "") if identity[key] else "" for key in ("grupo", "cota", "versao"))
            if key.strip("|"):
                existing[key] = row_number
        next_row = ws.max_row + 1
        for client in clients:
            values = client.variables or {}
            key = "|".join(str(values.get(item, "")) for item in ("grupo", "cota", "versao"))
            row_number = existing.get(key, next_row)
            if row_number == next_row:
                next_row += 1
            for field in fields:
                column = positions.get(field["internal_key"])
                if column:
                    ws.cell(row_number, column).value = values.get(field["internal_key"], "")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        ws.append([field["display_name"] for field in fields])
        for client in clients:
            values = client.variables or {}
            ws.append([values.get(field["internal_key"], "") for field in fields])
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def connector_service_account_email() -> str | None:
    raw = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if raw:
        try:
            return str(json.loads(raw).get("client_email") or "") or None
        except json.JSONDecodeError:
            return None
    path = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "").strip()
    if path:
        try:
            return str(json.loads(Path(path).read_text()).get("client_email") or "") or None
        except (OSError, json.JSONDecodeError):
            return None
    return None


def google_sheet_id(value: str) -> str:
    text = str(value or "").strip()
    match = re.search(r"/spreadsheets/d/([A-Za-z0-9_-]+)", text)
    return match.group(1) if match else text


def _service_account() -> dict[str, Any]:
    raw = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    path = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "").strip()
    if raw:
        return json.loads(raw)
    if path:
        return json.loads(Path(path).read_text())
    raise SystemSpreadsheetError("Service Account Google não configurada no servidor.")


def _google_token(*, readonly: bool) -> str:
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import padding
    import requests
    account = _service_account()
    encode = lambda value: base64.urlsafe_b64encode(json.dumps(value, separators=(",", ":")).encode()).decode().rstrip("=")
    now = int(time.time())
    header = encode({"alg": "RS256", "typ": "JWT"})
    claims = {"iss": account["client_email"], "scope": "https://www.googleapis.com/auth/spreadsheets.readonly" if readonly else "https://www.googleapis.com/auth/spreadsheets", "aud": "https://oauth2.googleapis.com/token", "iat": now, "exp": now + 3600}
    body = encode(claims)
    key = serialization.load_pem_private_key(account["private_key"].encode(), password=None)
    signature = key.sign(f"{header}.{body}".encode(), padding.PKCS1v15(), hashes.SHA256())
    assertion = f"{header}.{body}.{base64.urlsafe_b64encode(signature).decode().rstrip('=')}"
    response = requests.post("https://oauth2.googleapis.com/token", data={"grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer", "assertion": assertion}, timeout=20)
    if not response.ok:
        raise SystemSpreadsheetError("Não foi possível autenticar o conector Google.")
    return str(response.json().get("access_token") or "")


def _google_request(method: str, url: str, *, readonly: bool, payload: dict[str, Any] | None = None) -> dict[str, Any]:
    import requests
    try:
        response = requests.request(method, url, headers={"Authorization": f"Bearer {_google_token(readonly=readonly)}"}, json=payload, timeout=30)
    except requests.RequestException as exc:
        raise SystemSpreadsheetError("Falha de comunicação com Google Sheets.") from exc
    if not response.ok:
        raise SystemSpreadsheetError("Google Sheets recusou a operação.")
    return response.json() if response.content else {}


def test_google_connection(url_or_id: str) -> dict[str, Any]:
    sheet_id = google_sheet_id(url_or_id)
    if not sheet_id:
        raise SystemSpreadsheetError("Informe a URL ou ID da planilha Google.")
    try:
        metadata = _google_request("GET", f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}", readonly=True)
    except Exception as exc:
        raise SystemSpreadsheetError("Não foi possível acessar a planilha Google compartilhada.") from exc
    return {"spreadsheet_id": sheet_id, "name": metadata.get("properties", {}).get("title", ""), "tabs": [item.get("properties", {}).get("title", "") for item in metadata.get("sheets", [])], "service_account_email": connector_service_account_email()}


def import_google(*, name: str, url_or_id: str, tab: str, list_id: str | None = None, tenant_id: str = "default") -> dict[str, Any]:
    sheet_id = google_sheet_id(url_or_id)
    if not sheet_id or not tab:
        raise SystemSpreadsheetError("Informe a planilha e a aba Google.")
    try:
        values = _google_request("GET", f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{quote(tab, safe='')}", readonly=True).get("values", [])
    except SystemSpreadsheetError:
        raise
    except Exception as exc:
        raise SystemSpreadsheetError("Não foi possível ler os dados da aba Google.") from exc
    if not values:
        raise SystemSpreadsheetError("A aba Google não possui dados.")
    headers = [str(value or "").strip() for value in values[0]]
    result = create_system_spreadsheet(name, headers, list_id=list_id, tenant_id=tenant_id)
    with SessionLocal.begin() as db:
        sheet = _sheet(db, result["id"])
        _upsert_connector(db, sheet.id, "google_sheets", {"spreadsheet_id": sheet_id, "tab": tab})
        _upsert_rows(db, sheet.id, headers, values[1:], list_id=list_id)
    return get_system_spreadsheet(result["id"], tenant_id=tenant_id)


def attach_google(*, sheet_id: str, url_or_id: str, tab: str, tenant_id: str = "default") -> dict[str, Any]:
    get_system_spreadsheet(sheet_id, tenant_id=tenant_id)
    spreadsheet_id = google_sheet_id(url_or_id)
    if not spreadsheet_id or not tab:
        raise SystemSpreadsheetError("Informe a planilha e a aba Google.")
    values = _google_request("GET", f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{quote(tab, safe='')}", readonly=True).get("values", [])
    if not values:
        raise SystemSpreadsheetError("A aba Google não possui dados.")
    headers = [str(value or "").strip() for value in values[0]]
    with SessionLocal.begin() as db:
        _upsert_fields(db, _sheet(db, sheet_id), headers)
        _upsert_connector(db, sheet_id, "google_sheets", {"spreadsheet_id": spreadsheet_id, "tab": tab})
        _upsert_rows(db, sheet_id, headers, values[1:])
    return get_system_spreadsheet(sheet_id, tenant_id=tenant_id)


def sync_google(sheet_id: str, *, direction: str = "inbound", tenant_id: str = "default") -> dict[str, Any]:
    sheet = get_system_spreadsheet(sheet_id, tenant_id=tenant_id)
    connector_data: dict[str, Any] | None = None
    with SessionLocal() as db:
        connector = db.scalar(select(SpreadsheetConnector).where(SpreadsheetConnector.spreadsheet_id == sheet_id, SpreadsheetConnector.connector_type == "google_sheets"))
        if connector is None:
            raise SystemSpreadsheetError("A planilha não possui conector Google Sheets.")
        connector_data = dict(connector.configuration or {})
    spreadsheet_id = str(connector_data.get("spreadsheet_id") or "")
    tab = str(connector_data.get("tab") or "")
    if not spreadsheet_id or not tab:
        raise SystemSpreadsheetError("Conector Google incompleto.")
    try:
        if direction == "inbound":
            values = _google_request("GET", f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{quote(tab, safe='')}", readonly=True).get("values", [])
            if not values:
                raise SystemSpreadsheetError("A aba Google não possui dados.")
            with SessionLocal.begin() as db:
                _upsert_rows(db, sheet_id, [str(value or "") for value in values[0]], values[1:])
        elif direction == "outbound":
            fields = sheet["fields"]
            with SessionLocal() as db:
                clients = list(db.scalars(select(Client).where(Client.system_spreadsheet_id == sheet_id).order_by(Client.name)))
            values = [[field["display_name"] for field in fields]]
            values.extend([[str((client.variables or {}).get(field["internal_key"], "")) for field in fields] for client in clients])
            _google_request("PUT", f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{quote(tab, safe='')}?valueInputOption=RAW", readonly=False, payload={"range": tab, "majorDimension": "ROWS", "values": values})
        else:
            raise SystemSpreadsheetError("Direção de sincronização inválida.")
    except Exception as exc:
        with SessionLocal.begin() as db:
            connector = db.scalar(select(SpreadsheetConnector).where(SpreadsheetConnector.spreadsheet_id == sheet_id, SpreadsheetConnector.connector_type == "google_sheets"))
            if connector:
                connector.status, connector.last_error = "error", str(exc)[:500]
        if isinstance(exc, SystemSpreadsheetError):
            raise
        raise SystemSpreadsheetError("Sincronização Google falhou; os dados internos foram preservados.") from exc
    with SessionLocal.begin() as db:
        connector = db.scalar(select(SpreadsheetConnector).where(SpreadsheetConnector.spreadsheet_id == sheet_id, SpreadsheetConnector.connector_type == "google_sheets"))
        if connector:
            connector.status, connector.last_error, connector.last_synced_at = "synchronized", None, datetime.now(UTC)
    return get_system_spreadsheet(sheet_id, tenant_id=tenant_id)


def apply_action_outputs_to_system_spreadsheet(*, run_id: str, action_id: str, client_id: str | None, variables: dict[str, Any], result_payload: dict[str, Any], outputs: list[dict[str, Any]]) -> dict[str, Any]:
    if result_payload.get("_system_sheet_outputs_applied"):
        return {"applied": list(result_payload.get("system_sheet_outputs") or []), "synchronized_sheets": []}
    extracted = result_payload.get("dados_extraidos") if isinstance(result_payload.get("dados_extraidos"), dict) else {}
    applied: list[dict[str, Any]] = []
    sheet_ids: set[str] = set()
    with SessionLocal.begin() as db:
        client = db.get(Client, client_id) if client_id else None
        if client is None:
            grupo, cota, versao = (str(variables.get(key) or "") for key in ("grupo", "cota", "versao"))
            if grupo and cota:
                client = db.scalar(select(Client).where(Client.grupo == grupo, Client.cota == cota, Client.versao == versao).limit(1))
        for output in outputs:
            if not isinstance(output, dict):
                continue
            destination = output.get("destination") if isinstance(output.get("destination"), dict) else {}
            if destination.get("type") not in {"system_sheet_field", "data_source_field"}:
                continue
            field_id = str(destination.get("field_id") or "").strip()
            sheet_id = str(destination.get("system_spreadsheet_id") or destination.get("data_source_id") or "").strip()
            if not field_id or not sheet_id:
                raise SystemSpreadsheetError("Destino de output sem Planilha do Sistema ou field_id.")
            sheet = _sheet(db, sheet_id)
            field = db.get(DataSourceField, field_id)
            if field is None or field.data_source_id != sheet.id or not field.active:
                raise SystemSpreadsheetError("Output referencia um campo inexistente na Planilha do Sistema.")
            if client is None or client.system_spreadsheet_id != sheet.id:
                raise SystemSpreadsheetError("Cliente não pertence à Planilha do Sistema do output.")
            keys = {str(output.get("output_id") or ""), str(output.get("key") or ""), str(output.get("label") or ""), str(output.get("target_name") or "")}
            value = next((extracted[key] for key in extracted if str(key) in keys), None)
            if value is None and len(extracted) == 1:
                value = next(iter(extracted.values()))
            if value is None:
                continue
            variables_copy = dict(client.variables or {})
            internal_key = field.semantic_role or _key(field.display_name)
            variables_copy[internal_key] = str(value)
            client.variables = variables_copy
            if internal_key == "grupo": client.grupo = str(value)
            if internal_key == "cota": client.cota = str(value)
            if internal_key == "versao": client.versao = str(value)
            client.system_spreadsheet_id = sheet.id
            sheet_ids.add(sheet.id)
            applied.append({"output_id": output.get("output_id"), "field_id": field.id, "value": str(value), "system_spreadsheet_id": sheet.id})
        run = db.get(Run, run_id)
        if run is not None:
            diagnostics = dict(run.diagnostics or {})
            diagnostics["system_sheet_outputs"] = applied
            run.diagnostics = diagnostics
        result_payload["system_sheet_outputs"] = applied
        result_payload["_system_sheet_outputs_applied"] = True
    for sheet_id in sheet_ids:
        with SessionLocal() as db:
            connector = db.scalar(select(SpreadsheetConnector).where(SpreadsheetConnector.spreadsheet_id == sheet_id, SpreadsheetConnector.connector_type == "google_sheets"))
        if connector is not None:
            try:
                sync_google(sheet_id, direction="outbound")
            except SystemSpreadsheetError:
                pass
    return {"applied": applied, "synchronized_sheets": sorted(sheet_ids)}
